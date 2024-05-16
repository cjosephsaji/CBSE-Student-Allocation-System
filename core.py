from flask import Flask, render_template, request, send_file
import pandas as pd
import tabula
from openpyxl import Workbook

app = Flask(__name__)

school  = 'SCHOOL NAME'

subject_codes = {
    "016": "ARABIC",
    "018": "FRENCH",
    "041": "MATHEMATICS",
    "085": "HINDI COURSE-B",
    "086": "SCIENCE THEORY",
    "087": "SOCIAL SCIENCE",
    "119": "SANSKRIT",
    "184": "ENGLISH LNG & LIT.",
    "241": "BASIC MATHEMATICS",
    "303": "URDU COURSE - B",
    "402": "INFO TECHNOLOGY",
    "027": "HISTORY",
    "030": "ECONOMICS",
    "034": "MUSIC HINDUSTANI (VOCAL)",
    "037": "PSYCHOLOGY",
    "042": "PHYSICS",
    "043": "CHEMISTRY",
    "044": "BIOLOGY",
    "048": "PHYSICAL EDUCATION",
    "054": "BUSINESS STUDIES",
    "064": "HOME SCIENCE",
    "065": "INFORMATICS PRACTICE",
    "066": "ENTREPRENEURSHIP",
    "083": "COMPUTER SCIENCE",
    "301": "ENGLISH CORE",
    "802": "INFORMATION TECHNOLOGY",
    "806": "TOURISM",
    "812": "MARKETING",
    "834": "FOOD SCIENCE AND NUTRITION"
}

@app.route('/', methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        f = request.files['the_file']
        if f.filename[-4:] == '.pdf':
            subjects = request.form['subjects'].split(',')
            number_of_rows = int(request.form['rows_number'])
            number_of_cols = int(request.form['cols_number'])
            number_of_students_per_class = request.form['number_of_students_per_class']
            tabula.convert_into(f, "output.csv", output_format="csv", pages='all', lattice=True)
            f = pd.read_csv('output.csv', on_bad_lines="skip")
            roll_number_col = 1
            dict_students = {}
            for i,row in f.iterrows():
             if row['ROLL-NO'] != 'ROLL-NO':
                for subject in subjects:
                    subject = subject.strip()
                    try:
                        if str(row[subject]).zfill(3) == str(subject).zfill(3):
                            if subject in dict_students:
                                for col in row:
                                    col = str(col)
                                    if col != '...' and len(col) > 3:
                                        dict_students[subject].append(col)
                            else:
                                for col in row:
                                    col = str(col)
                                    if col != '...' and len(col) > 3:
                                        dict_students[subject] = []
                                        dict_students[subject].append(col)
                    except KeyError:
                        return render_template('home.html', error='Please check whether the entered subject codes correspond to the pdf uploaded')

            wb = Workbook()
            for subject in subjects:
                subject = subject.strip()
                start,stop=0,int(number_of_students_per_class)
                count = 0
                ws = wb.create_sheet(subject)
                ws = wb.create_sheet(subject + " ROOM PLAN")
                ws.append(['ROLL NO', 'ROOM NUMBER'])
                excel_row = 8
                while True:
                    try:
                        students = sorted(dict_students[subject])[start:stop]
                        if students != []:
                            count +=1
                            roll_number_range = f'{students[0]} - {students[-1]}'
                            ws2 = wb[subject]
                            ws2.append([f'NAME OF CENTRE : {school}', '', '', '', f'ROOM NUMBER : {count}'])
                            ws2.append([f'NAME OF THE EXAMINATION : '])
                            ws2.append([f'SUBJECT/CODE :  {subject_codes[subject]}/{subject} '])
                            ws2.append([f'DAY/DATE : ', '', '','','CENTRE No. : '])
                            rows,columns = number_of_rows,number_of_cols
                            student_index = 0
                            total_registered_students = 0
                            for c in range(columns):
                                for r in range(rows):
                                    try:
                                        ws2.cell(row=r+excel_row, column=c+1).value = students[student_index]
                                        total_registered_students += 1
                                    except IndexError:
                                        ws2.cell(row=r+excel_row, column=c+1).value = ''
                                    student_index += 1

                            ws2.append([''])
                            ws2.append(['SIGNATURE OF ASST. SUPERINTENDENTS', '', '', '', f'TOTAL NUMBER REGISTERED : {total_registered_students}'])
                            ws2.append([''])
                            ws2.append(['1.......................', 'NO : PRESENT..............................', 'NO : ABSENT..............................'])
                            ws2.append(['2.......................'])
                            ws2.append(['ROLL NO. OF ABSENTEES : ...............................................'])
                            ws2.append([''])
                            ws2.append(['','','','','CENTRE SUPERINTENDENT'])

                            for _ in range(3):
                                ws2.append([''])
                            excel_row = excel_row + rows + 18
                            try:
                                ws.append([roll_number_range, count])
                            except IndexError:
                                ws.append([roll_number_range, count])
                            start=stop
                            stop += int(number_of_students_per_class)
                            if stop > len(dict_students[subject]):
                                stop = len(dict_students[subject])
                        else:
                            break
                    except KeyError:
                        print(f'No student opted for {subject}')
                        break
            wb.save('data.xlsx')
            return render_template('sorted_data.html', dict_students=dict_students)
        else:
            return render_template('home.html', error='Please upload a PDF file ONLY')
    else:
        return render_template('home.html')


@app.route('/excelstudentdata')
def download_excel_student_data():
    file = "data.xlsx"
    return send_file(file,as_attachment=True)
